import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from datetime import datetime

def exportMatchedToExcel():
    """Reads the JSON file and exports matched articles to an Excel file with formatted dates and clickable hyperlinks."""
    with open('newsData.json', 'r', encoding='utf-8') as file:
        newsData = json.load(file)
    
    # Filter only matched=True articles
    matchedArticles = [article for article in newsData if article.get('matched')]
    
    # Sort by date (assuming date is in YYYY-MM-DD format)
    matchedArticles.sort(key=lambda x: x['date'])
    
    # Prepare data for DataFrame
    data = []
    for idx, article in enumerate(matchedArticles, start=1):
        # Convert date to "13 May 2024" format
        formattedDate = datetime.strptime(article['date'], "%Y-%m-%d").strftime("%d %B %Y")
        data.append([idx, formattedDate, article['title'], article['link'], ' '])  # Added empty remarks column
    
    # Create DataFrame
    df = pd.DataFrame(data, columns=['Sr No.', 'Date', 'Title', 'URL', 'Remarks'])  # Added Remarks column
    
    # Save to Excel
    excelFilename = "DataCentreNEWS.xlsx"
    with pd.ExcelWriter(excelFilename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    
    # Load the workbook and worksheet
    workbook = load_workbook(excelFilename)
    worksheet = workbook.active

    # Adjust column widths dynamically
    worksheet.column_dimensions[get_column_letter(2)].width = 15  # Fit date column width
    worksheet.column_dimensions[get_column_letter(3)].width = 50  # Title column width
    worksheet.column_dimensions[get_column_letter(4)].width = 50  # URL column width
    worksheet.column_dimensions[get_column_letter(5)].width = 30  # Added Remarks column width

    # Apply hyperlink formatting
    hyperlinkFont = Font(color="0000FF", underline="single")  # Blue font with underline
    for row in range(2, len(matchedArticles) + 2):  # Start from row 2 (row 1 is headers)
        cell = worksheet.cell(row=row, column=4)  # URL column
        cell.hyperlink = cell.value  # Set hyperlink
        cell.font = hyperlinkFont   # Apply styling

    # Save the modified workbook
    workbook.save(excelFilename)
    
    print(f"Successfully saved {len(matchedArticles)} matched articles to {excelFilename} with formatted dates and clickable links.")

if __name__ == "__main__":
    exportMatchedToExcel()
